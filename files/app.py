import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

API_URL = "http://localhost:8000/generate"  # vLLM sunucun

# Çeviri fonksiyonu
def translate_text(text):
    payload = {
        "prompt": f"Translate to Turkish:\n{text}",
        "max_tokens": 1000,
        "temperature": 0.7,
    }
    try:
        response = requests.post(API_URL, json=payload, timeout=10)
        response.raise_for_status()
        return response.json()["text"].strip()
    except Exception as e:
        print(f"Failed to translate '{text}': {e}")
        return text  # hata varsa orijinal metni bırak

# Ana işlem fonksiyonu
def translate_excel_parallel(input_path, output_path, max_workers=8):
    wb = openpyxl.load_workbook(input_path)
    sheet = wb.active  # İstersen tüm sheet'leri dönebilirsin

    # Çevrilecek hücrelerin koordinat ve içeriklerini topla
    translatable_cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                translatable_cells.append(((cell.row, cell.column), cell.value))

    print(f"Found {len(translatable_cells)} text cells to translate.")

    # Çevirileri paralel olarak al
    translated_texts = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_cell = {
            executor.submit(translate_text, text): pos
            for pos, text in translatable_cells
        }
        for future in as_completed(future_to_cell):
            pos = future_to_cell[future]
            try:
                translated = future.result()
                translated_texts[pos] = translated
            except Exception as e:
                print(f"Error at {pos}: {e}")
                translated_texts[pos] = None

    # Çevrilmiş metinleri hücrelere yerleştir
    for (row, col), translated in translated_texts.items():
        if translated:
            sheet.cell(row=row, column=col).value = translated

    wb.save(output_path)
    print(f"Translated Excel saved as: {output_path}")
